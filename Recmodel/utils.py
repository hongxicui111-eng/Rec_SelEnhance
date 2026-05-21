# -*- coding: utf-8 -*-
# @Time    : 2020/3/30 11:06
# @Author  : Hui Wang

import numpy as np
import math
import random
import os
import json
import pickle
from scipy.sparse import csr_matrix
from texttable import Texttable
import torch
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import openpyxl
import torch.nn.functional as F
import pandas as pd
import ast
from collections import Counter
from sklearn.metrics import auc as sk_auc


def set_seed(seed):
    random.seed(seed)
    os.environ['PYTHONHASHSEED'] = str(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    # some cudnn methods can be random even after fixing the seed
    # unless you tell it to be deterministic
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False


def check_path(path):
    if not os.path.exists(path):
        os.makedirs(path)
        print(f'{path} created')

def neg_sample(target, item_size):
    item = random.randint(1, item_size - 1)
    while item == target:
        item = random.randint(1, item_size - 1)
    return item

def generate_scaled_fx(k=0.03, size=200):
    """
    生成缩放后的函数值 fx 和其总和。

    参数:
    k -- 控制下降速度的常数
    size -- 数据的大小，默认为 101

    返回:
    fx -- 缩放后的函数值
    total_sum -- fx 的总和
    """
    x = np.arange(0, size)
    fx_raw = np.exp(-k * x)

    # 缩放因子 A，使总和为 0.99
    A = 0.99 / np.sum(fx_raw)
    fx = A * fx_raw

    s = [0]
    temps = 0
    for i in range(len(fx)):
        temps += fx[i]
        s.append(temps) 

    
    return fx, s


def neg_sample_dns(target, item_sequence, item_size):
    """
    从 [1, item_size-1] 中采样一个不等于 target 且不在 item_sequence 中的负样本。
    
    :param target: 当前的正样本
    :param item_sequence: 用户历史交互的物品列表
    :param item_size: 物品总数
    :return: 一个负样本的物品 ID
    """
    item = random.randint(1, item_size - 1)
    while item == target or item in item_sequence:
        item = random.randint(1, item_size - 1)
    return item



def neg_sample_dns_unique(target, item_sequence, item_size, N):
    """
    采样 N 个不等于 target 且不在 item_sequence 中的互不相同的负样本。
    
    :param target: 当前的正样本
    :param item_sequence: 用户历史交互的物品列表
    :param item_size: 物品总数
    :param N: 需要采样的负样本数
    :return: List[int]，长度为 N 的负样本 ID 列表
    """
    item_sequence_set = set(item_sequence)
    invalid_items = item_sequence_set.union(set(target))
    
    # 所有合法的负样本集合
    all_neg_candidates = list(set(range(1, item_size)) - invalid_items)

    if len(all_neg_candidates) < N:
        raise ValueError(f"可选的负样本数不足：需要 {N} 个，只有 {len(all_neg_candidates)} 个可选")

    return random.sample(all_neg_candidates, N)


class EarlyStopping:
    """Early stops the training if validation loss doesn't improve after a given patience."""
    def __init__(self, checkpoint_path, patience=7, verbose=False, delta=0):
        """
        Args:
            patience (int): How long to wait after last time validation loss improved.
                            Default: 7
            verbose (bool): If True, prints a message for each validation loss improvement.
                            Default: False
            delta (float): Minimum change in the monitored quantity to qualify as an improvement.
                            Default: 0
        """
        self.checkpoint_path = checkpoint_path
        self.patience = patience
        self.verbose = verbose
        self.counter = 0
        self.best_score = None
        self.early_stop = False
        self.delta = delta

    def compare(self, score):
        for i in range(len(score)):
            # 有一个指标增加了就认为是还在涨
            if score[i] > self.best_score[i]+self.delta:
                return False
        return True

    def __call__(self, score, model):
        # score HIT@10 NDCG@10

        if self.best_score is None:
            self.best_score = score
            self.score_min = np.array([0]*len(score))
            self.save_checkpoint(score, model)
        elif self.compare(score):
            self.counter += 1
            print(f'EarlyStopping counter: {self.counter} out of {self.patience}')
            if self.counter >= self.patience:
                self.early_stop = True
        else:
            self.best_score = score
            self.save_checkpoint(score, model)
            self.counter = 0

    def save_checkpoint(self, score, model):
        '''Saves model when validation loss decrease.'''
        if self.verbose:
            # ({self.score_min:.6f} --> {score:.6f}) # 这里如果是一个值的话输出才不会有问题
            print(f'Validation score increased.  Saving model ...')
        torch.save(model.state_dict(), self.checkpoint_path)
        self.score_min = score

def kmax_pooling(x, dim, k):
    index = x.topk(k, dim=dim)[1].sort(dim=dim)[0]
    return x.gather(dim, index).squeeze(dim)

def avg_pooling(x, dim):
    return x.sum(dim=dim)/x.size(dim)


def generate_rating_matrix_valid(user_seq, num_users, num_items):
    # three lists are used to construct sparse matrix
    row = []
    col = []
    data = []
    for user_id, item_list in enumerate(user_seq):
        for item in item_list[:-2]: #
            row.append(user_id)
            col.append(item)
            data.append(1)

    row = np.array(row)
    col = np.array(col)
    data = np.array(data)
    rating_matrix = csr_matrix((data, (row, col)), shape=(num_users, num_items))

    return rating_matrix

def generate_rating_matrix_test(user_seq, num_users, num_items):
    # three lists are used to construct sparse matrix
    row = []
    col = []
    data = []
    for user_id, item_list in enumerate(user_seq):
        for item in item_list[:-1]: #
            row.append(user_id)
            col.append(item)
            data.append(1)

    row = np.array(row)
    col = np.array(col)
    data = np.array(data)
    rating_matrix = csr_matrix((data, (row, col)), shape=(num_users, num_items))

    return rating_matrix

def get_user_seqs(data_file):
    item_counter = Counter()
    lines = open(data_file).readlines()
    user_seq = []
    item_set = set()
    user_index = 0
    for line in tqdm(lines):
        items = line.strip().split(' ')
        items = [int(item) for item in items[:-1]] #最后一个是时间戳
        user_seq.append(items)
        item_set = item_set | set(items)
        item_counter.update(items)
    max_item = max(item_set)

    num_users = len(lines)
    num_items = max_item + 2
    return user_seq, max_item, item_counter

def get_user_seqs_long(data_file):
    lines = open(data_file).readlines()
    user_seq = []
    long_sequence = []
    item_set = set()
    for line in lines:
        user, items = line.strip().split(' ', 1)
        items = items.split(' ')
        items = [int(item) for item in items]
        long_sequence.extend(items) # 后面的都是采的负例
        user_seq.append(items)
        item_set = item_set | set(items)
    max_item = max(item_set)

    return user_seq, max_item, long_sequence

def get_user_seqs_and_sample(data_file, sample_file):
    lines = open(data_file).readlines()
    user_seq = []
    item_set = set()
    for line in lines:
        user, items = line.strip().split(' ', 1)
        items = items.split(' ')
        items = [int(item) for item in items]
        user_seq.append(items)
        item_set = item_set | set(items)
    max_item = max(item_set)

    lines = open(sample_file).readlines()
    sample_seq = []
    for line in lines:
        user, items = line.strip().split(' ', 1)
        items = items.split(' ')
        items = [int(item) for item in items]
        sample_seq.append(items)

    assert len(user_seq) == len(sample_seq)

    return user_seq, max_item, sample_seq

def get_item2attribute_json(data_file):
    item2attribute = json.loads(open(data_file).readline())
    attribute_set = set()
    for item, attributes in item2attribute.items():
        attribute_set = attribute_set | set(attributes)
    attribute_size = max(attribute_set) # 331
    return item2attribute, attribute_size

def get_metric(pred_list, topk=10):
    NDCG = 0.0
    HIT = 0.0
    MRR = 0.0
    # [batch] the answer's rank
    for rank in pred_list:
        if rank < topk:
            MRR += 1.0 / (rank + 1.0)
            NDCG += 1.0 / np.log2(rank + 2.0)
            HIT += 1.0
    return HIT /len(pred_list), NDCG /len(pred_list), MRR /len(pred_list)

def get_metric_stage2(pred_list, topk=10):
    logloss = 0.0
    auc = 0.0
    # [batch] the answer's rank
    
    for rank in pred_list:
        if rank < topk:
            MRR += 1.0 / (rank + 1.0)
            NDCG += 1.0 / np.log2(rank + 2.0)
            HIT += 1.0
    return LogLoss(logloss), AUC(auc)

def LogLoss(preds, trues):
    r"""Logloss_ (also known as logistic loss or cross-entropy loss) is used to evaluate the probabilistic
    output of the two-class classifier.

    .. _Logloss: http://wiki.fast.ai/index.php/Log_Loss

    .. math::
        LogLoss = \frac{1}{|S|} \sum_{(u,i) \in S}(-((r_{u i} \ \log{\hat{r}_{u i}}) + {(1 - r_{u i})}\ \log{(1 - \hat{r}_{u i})}))
    """
    eps = 1e-15
    preds = np.float64(preds)
    preds = np.clip(preds, eps, 1 - eps)
    loss = np.sum(-trues * np.log(preds) - (1 - trues) * np.log(1 - preds))
    return loss / len(preds)
    
def _binary_clf_curve(trues, preds):
    """Calculate true and false positives per binary classification threshold

    Args:
        trues (numpy.ndarray): the true scores' list
        preds (numpy.ndarray): the predict scores' list

    Returns:
        fps (numpy.ndarray): A count of false positives, at index i being the number of negative
        samples assigned a score >= thresholds[i]
        preds (numpy.ndarray): An increasing count of true positives, at index i being the number
        of positive samples assigned a score >= thresholds[i].

    Note:
        To improve efficiency, we referred to the source code(which is available at sklearn.metrics.roc_curve)
        in SkLearn and made some optimizations.

    """
    trues = trues == 1

    desc_idxs = np.argsort(preds)[::-1]
    preds = preds[desc_idxs]
    trues = trues[desc_idxs]

    unique_val_idxs = np.where(np.diff(preds))[0]
    threshold_idxs = np.r_[unique_val_idxs, trues.size - 1]

    tps = np.cumsum(trues)[threshold_idxs]
    fps = 1 + threshold_idxs - tps
    return fps, tps

def AUC(preds, trues):
    r"""AUC_ (also known as Area Under Curve) is used to evaluate the two-class model, referring to
    the area under the ROC curve.

    .. _AUC: https://en.wikipedia.org/wiki/Receiver_operating_characteristic#Area_under_the_curve

    Note:
        This metric does not calculate group-based AUC which considers the AUC scores
        averaged across users. It is also not limited to k. Instead, it calculates the
        scores on the entire prediction results regardless the users. We call the interface
        in `scikit-learn`, and code calculates the metric using the variation of following formula.

    .. math::
        \mathrm {AUC} = \frac {{{M} \times {(N+1)} - \frac{M \times (M+1)}{2}} -
        \sum\limits_{i=1}^{M} rank_{i}} {{M} \times {(N - M)}}

    :math:`M` denotes the number of positive items.
    :math:`N` denotes the total number of user-item interactions.
    :math:`rank_i` denotes the descending rank of the i-th positive item.
    """

    fps, tps = _binary_clf_curve(trues, preds)
    if len(fps) > 2:
        optimal_idxs = np.where(
            np.r_[True, np.logical_or(np.diff(fps, 2), np.diff(tps, 2)), True]
        )[0]
        fps = fps[optimal_idxs]
        tps = tps[optimal_idxs]

    tps = np.r_[0, tps]
    fps = np.r_[0, fps]

    if fps[-1] <= 0:
        logger = getLogger()
        logger.warning(
            "No negative samples in y_true, "
            "false positive value should be meaningless"
        )
        fpr = np.repeat(np.nan, fps.shape)
    else:
        fpr = fps / fps[-1]

    if tps[-1] <= 0:
        logger = getLogger()
        logger.warning(
            "No positive samples in y_true, "
            "true positive value should be meaningless"
        )
        tpr = np.repeat(np.nan, tps.shape)
    else:
        tpr = tps / tps[-1]

    result = sk_auc(fpr, tpr)
    return result

def precision_at_k_per_sample(actual, predicted, topk):
    num_hits = 0
    for place in predicted:
        if place in actual:
            num_hits += 1
    return num_hits / (topk + 0.0)

def precision_at_k(actual, predicted, topk):
    sum_precision = 0.0
    num_users = len(predicted)
    for i in range(num_users):
        act_set = set(actual[i])
        pred_set = set(predicted[i][:topk])
        sum_precision += len(act_set & pred_set) / float(topk)

    return sum_precision / num_users

def recall_at_k(actual, predicted, topk):
    sum_recall = 0.0
    num_users = len(predicted)
    true_users = 0
    for i in range(num_users):
        act_set = set(actual[i])
        pred_set = set(predicted[i][:topk])
        if len(act_set) != 0:
            sum_recall += len(act_set & pred_set) / float(len(act_set))
            true_users += 1
    return sum_recall / true_users


def apk(actual, predicted, k=10):
    """
    Computes the average precision at k.
    This function computes the average precision at k between two lists of
    items.
    Parameters
    ----------
    actual : list
             A list of elements that are to be predicted (order doesn't matter)
    predicted : list
                A list of predicted elements (order does matter)
    k : int, optional
        The maximum number of predicted elements
    Returns
    -------
    score : double
            The average precision at k over the input lists
    """
    if len(predicted)>k:
        predicted = predicted[:k]

    score = 0.0
    num_hits = 0.0

    for i,p in enumerate(predicted):
        if p in actual and p not in predicted[:i]:
            num_hits += 1.0
            score += num_hits / (i+1.0)

    if not actual:
        return 0.0

    return score / min(len(actual), k)


def mapk(actual, predicted, k=10):
    """
    Computes the mean average precision at k.
    This function computes the mean average prescision at k between two lists
    of lists of items.
    Parameters
    ----------
    actual : list
             A list of lists of elements that are to be predicted
             (order doesn't matter in the lists)
    predicted : list
                A list of lists of predicted elements
                (order matters in the lists)
    k : int, optional
        The maximum number of predicted elements
    Returns
    -------
    score : double
            The mean average precision at k over the input lists
    """
    return np.mean([apk(a, p, k) for a, p in zip(actual, predicted)])

def ndcg_k(actual, predicted, topk):
    res = 0
    for user_id in range(len(actual)):
        k = min(topk, len(actual[user_id]))
        idcg = idcg_k(k)
        dcg_k = sum([int(predicted[user_id][j] in
                         set(actual[user_id])) / math.log(j+2, 2) for j in range(topk)])
        res += dcg_k / idcg
    return res / float(len(actual))


# Calculates the ideal discounted cumulative gain at k
def idcg_k(k):
    res = sum([1.0/math.log(i+2, 2) for i in range(k)])
    if not res:
        return 1.0
    else:
        return res

def generate_rating_matrix(user_seq, num_items):
    # three lists are used to construct sparse matrix
    row = []
    col = []
    data = []
    num_users = len(user_seq)
    for user_id, item_list in enumerate(user_seq):
        for item in item_list[:-1]:  #
            row.append(user_id)
            col.append(item)
            data.append(1)

    row = np.array(row)
    col = np.array(col)
    data = np.array(data)
    rating_matrix = csr_matrix((data, (row, col)), shape=(num_users, num_items))

    return rating_matrix

def get_gpu_usage(device=None):
    r"""Return the reserved memory and total memory of given device in a string.
    Args:
        device: cuda.device. It is the device that the model run on.

    Returns:
        str: it contains the info about reserved memory and total memory of given device.
    """

    reserved = torch.cuda.max_memory_reserved(device) / 1024**3
    total = torch.cuda.get_device_properties(device).total_memory / 1024**3

    return "{:.2f} G/{:.2f} G".format(reserved, total)


def get_environment():
    device = "cuda"
    # print(device)
    gpu_usage = get_gpu_usage(device)

    import psutil

    memory_used = psutil.Process(os.getpid()).memory_info().rss / 1024**3
    memory_total = psutil.virtual_memory()[0] / 1024**3
    memory_usage = "{:.2f} G/{:.2f} G".format(memory_used, memory_total)
    cpu_usage = "{:.2f} %".format(psutil.cpu_percent(interval=1))
    """environment_data = [
        {"Environment": "CPU", "Usage": cpu_usage,},
        {"Environment": "GPU", "Usage": gpu_usage, },
        {"Environment": "Memory", "Usage": memory_usage, },
    ]"""

    table = Texttable()
    table.set_cols_align(["l", "c"])
    table.set_cols_valign(["m", "m"])
    table.add_rows(
        [
            ["Environment", "Usage"],
            ["CPU", cpu_usage],
            ["GPU", gpu_usage],
            ["Memory", memory_usage],
        ]
    )

    return table


import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl import load_workbook
import ast

def to_excel(result_info, args, stage, epoch, training_time, inference_time):
    # 转换result_info为字典，并将键中的下划线替换为@
    result_info = ast.literal_eval(result_info)
    result_info = {key.replace('_', '@'): value for key, value in result_info.items()}

    # 添加额外的信息
    result_info['Training_time'] = training_time
    result_info['Inference_time'] = inference_time
    result_info['GPU'] = get_gpu_usage("cuda")
    result_info['DataSet'] = args.data_name
    result_info['Model'] = args.backbone
    result_info['Stage'] = stage
    result_info['N'] = args.N
    result_info['M'] = args.M
    result_info['K'] = args.K
    result_info['CL'] = args.CL_type

    keys_to_remove = ['Epoch', 'MRR@5', 'MRR@10', 'MRR@20']
    for key in keys_to_remove:
        if key in result_info:
            del result_info[key]
    result_info['Epoch'] = epoch

    # 文件名
    file_name = 'Yelp_N_500.xlsx'
    print(file_name)
    
    try:
        # 读取现有的Excel文件
        df = pd.read_excel(file_name)
    except FileNotFoundError:
        # 如果文件不存在，则创建一个新的DataFrame
        df = pd.DataFrame(columns=['DataSet', 'Model', 'Stage', 'Epoch' ,'HIT@5', 'HIT@10', 'HIT@20', 'NDCG@5', 'NDCG@10', 'NDCG@20', 'Training_time', 'GPU', 'Inference_time'])

    # 将新的结果转换为DataFrame
    new_result_df = pd.DataFrame([result_info])

    # 将新的结果添加到现有的DataFrame中
    df = pd.concat([df, new_result_df], ignore_index=True)
    # 按照'DataSet'和'Model'列进行排序
    df = df.sort_values(by=['DataSet', 'Model', 'Stage'])

    # 先将数据保存至Excel文件
    df.to_excel(file_name, index=False)

    # 使用openpyxl打开已保存的文件进行样式处理
    workbook = openpyxl.load_workbook(file_name)

    # 遍历所有工作表
    for sheet in workbook.worksheets:
        # 遍历每一行
        for row in sheet.iter_rows():
            # 遍历每一个单元格
            for cell in row:
                # 设置单元格居中
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=False)

    # 保存文件
    workbook.save(file_name)

    # 加粗最大值
    workbook = load_workbook(file_name)
    sheet = workbook.active

    # 获取唯一的数据集值
    datasets = df['DataSet'].unique()

    # 定义需要处理的指标列
    metrics = ['HR@5', 'HR@10', 'HR@20', 'NDCG@5', 'NDCG@10', 'NDCG@20']

    for dataset in datasets:
        dataset_df = df[df['DataSet'] == dataset]

        # 对每一个指标列
        for metric in metrics:
            if metric in dataset_df.columns:
                if dataset_df[metric].dtype == 'object':
                    dataset_df[metric] = pd.to_numeric(dataset_df[metric])

                # 找到最大值
                max_value = dataset_df[metric].max()

                # 遍历所有模型，并找出最大值所在的行
                for index, row in dataset_df.iterrows():
                    if row[metric] == max_value:
                        cell = sheet.cell(row=index + 2, column=df.columns.get_loc(metric) + 1)
                        cell.font = Font(bold=True)

    # 保存文件
    workbook.save(file_name)




    
